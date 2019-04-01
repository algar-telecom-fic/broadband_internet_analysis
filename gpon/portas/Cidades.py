from Database import Database
from openpyxl import load_workbook


class Cidades:
    def __init__(self, filename):
        self.concessao = []
        self.expansao  = []

        self.processaCidades(filename)

    def processaCidades(self, filename):
        db = Database('dbconfigs.env')

        if filename == "":
            query = "SELECT nome, tipo FROM cidades"
            result = db.executaQuery(query)

            self.concessao = []
            self.expansao  = []

            for item in result:
                if item[1] == 'CONCESSÃO': self.concessao.append(item[0])
                else:                      self.expansao.append(item[0])

        else:
            wb = load_workbook(filename)
            ws = wb['todas_localidades_existentes']

            self.concessao = []
            self.expansao  = []

            insert_args = []
            for i in range(2, ws.max_row+1):

                localidade = ws.cell(row=i, column=1).value
                tipo       = ws.cell(row=i, column=2).value
                regional   = ws.cell(row=i, column=3).value

                if tipo == 'CONCESSÃO': self.concessao.append(localidade)
                else:                   self.expansao.append(localidade)

                insert_args.append( (localidade, tipo, regional) )

            query = """INSERT INTO cidades(nome, tipo, regional) VALUES(%s, %s, %s)
                       ON DUPLICATE KEY UPDATE
                       tipo = VALUES(tipo),
                       regional = VALUES(regional); """

            db.executaQuery(query, insert_args)
