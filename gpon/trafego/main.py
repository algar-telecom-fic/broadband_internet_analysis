from openpyxl import load_workbook
from find_ip import find_ip



def read_file(filename):
    wb = load_workbook(filename)
    ws = wb['GPON_2019-02-11']
    
    global database
    database = {}
    
    for i in range(2, ws.max_row):
        current_ip = find_ip( ws.cell(row = i, column = 7).value )
        
        if current_ip not in database:
            database[current_ip] = { 'capacidadeMax': 0,
                                     'utilizacaoGbps': 0,
                                     'utilizacaoPrcnt': 0, }
        
        current_capacity = int(ws.cell(row = i, column = 10).value)
        Gbps = (float(ws.cell(row = i, column = 11).value) * current_capacity) / 100
        
        database[current_ip]['capacidadeMax'] += current_capacity
        database[current_ip]['utilizacaoGbps'] += Gbps
        
    
    
    for ip in database:
        try:
            database[ip]['utilizacaoPrcnt'] = (database[ip]['utilizacaoGbps'] / database[ip]['capacidadeMax']) * 100
        except:
            database[ip]['utilizacaoPrcnt'] = 0.00
    
    print("IP\tCAPACIDADE_MAX\tUTILIZACAOgbps\tUTILIZACAOprcnt")
    
    for ip in database:
        print( '{}\t{}\t{}\t{:.0f}%'.format(
            ip,
            database[ip]['capacidadeMax'],
            database[ip]['utilizacaoGbps'],
            database[ip]['utilizacaoPrcnt'],
        ))

read_file('datasheets/GPON_2019-02-11.xlsx')
