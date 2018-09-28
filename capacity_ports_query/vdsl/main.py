def add_port(filename, regional, locale, station, cabinet, port):
	global database
	if regional not in database[filename]:
		database[filename][regional] = {}
	if locale not in database[filename][regional]:
		database[filename][regional][locale] = {}
	if station not in database[filename][regional][locale]:
		database[filename][regional][locale][station] = {}
	if cabinet not in database[filename][regional][locale][station]:
		database[filename][regional][locale][station][cabinet] = {
			'total': 0,
			'available': 0,
			'occupied': 0,
		}
	database[filename][regional][locale][station][cabinet]['total'] += 1
	if port > 0:
		database[filename][regional][locale][station][cabinet]['available'] += 1
	elif port < 0:
		database[filename][regional][locale][station][cabinet]['occupied'] += 1

def build_database(current_file):
	global database
	v = []
	for regional in database[current_file]:
			for locale in database[current_file][regional]:
				for station in database[current_file][regional][locale]:
					for cabinet in database[current_file][regional][locale][station]:
						v.append(
							data(
								regional,
								locale,
								station,
								cabinet,
								database[current_file][regional][locale][station][cabinet]['total'],
								database[current_file][regional][locale][station][cabinet]['available'],
								database[current_file][regional][locale][station][cabinet]['occupied'],
							)
						)
	v.sort()
	database[current_file] = v

def build_excel_file(current_file, previous_file, date_difference):
	global database, excel_file
	from math import ceil
	from openpyxl import Workbook
	from openpyxl.utils import get_column_letter
	from openpyxl.styles import PatternFill
	excel_file = Workbook()
	sheet = excel_file.worksheets[0]
	sheet.title = 'Results'
	build_styles()
	current_row = 1
	columns = 9
	greatest_column = [0 for i in range(columns + 1)]
	sheet.cell(row = current_row, column = 1).value = 'Regional'
	sheet.cell(row = current_row, column = 2).value = 'Localidade'
	sheet.cell(row = current_row, column = 3).value = 'Estação mãe'
	sheet.cell(row = current_row, column = 4).value = 'Armário'
	sheet.cell(row = current_row, column = 5).value = 'Total de portas'
	sheet.cell(row = current_row, column = 6).value = 'Portas disponíveis'
	sheet.cell(row = current_row, column = 7).value = 'Portas ocupadas'
	sheet.cell(row = current_row, column = 8).value = 'Crescimento'
	sheet.cell(row = current_row, column = 9).value = 'Previsão de esgotamento'
	for j in range(1, columns + 1):
		sheet.cell(row = current_row, column = j).style = 'top_style'
		greatest_column[j] = max(greatest_column[j], len(str(sheet.cell(row = current_row, column = j).value)))
	for i in database[current_file]:
		current_row += 1
		sheet.cell(row = current_row, column = 1).value = i.regional
		sheet.cell(row = current_row, column = 2).value = i.locale
		sheet.cell(row = current_row, column = 3).value = i.station
		sheet.cell(row = current_row, column = 4).value = i.cabinet
		sheet.cell(row = current_row, column = 5).value = i.total
		sheet.cell(row = current_row, column = 6).value = i.available
		sheet.cell(row = current_row, column = 7).value = i.occupied
		try:
			before = database[previous_file][i.regional][i.locale][i.station][i.cabinet]['occupied']
			median = (i.occupied - before) / (date_difference / 30)
			sheet.cell(row = current_row, column = 8).value = round(median, 1)
			if i.available == 0:
				sheet.cell(row = current_row, column = 9).value = 'Esgotado'
			elif median > 0:
				value = ceil(i.available / median)
				if value == 1:
					sheet.cell(row = current_row, column = 9).value = 'Esgota em até 1 mês'
				elif value > 10:
					sheet.cell(row = current_row, column = 9).value = 'Esgota em mais de 10 meses'
				else:
					sheet.cell(row = current_row, column = 9).value = 'Esgota em até ' + str(value) + ' meses'
			elif median < 0:
				sheet.cell(row = current_row, column = columns).value = 'Decrescimento'
			else:
				sheet.cell(row = current_row, column = columns).value = 'Estável'
		except Exception as e:
			print(e)
			sheet.cell(row = current_row, column = 8).value = 'Sem histórico'
			sheet.cell(row = current_row, column = 9).value = 'Sem histórico'
		for j in range(1, columns + 1):
			sheet.cell(row = current_row, column = j).style = 'normal_style'
			greatest_column[j] = max(greatest_column[j], len(str(sheet.cell(row = current_row, column = j).value)))
		current_color = 'FF0000' if i.available == 0 else ('00FF00' if i.available > 10 else 'FFFF00')
		sheet.cell(row = current_row, column = 6).fill = PatternFill('solid', fgColor = current_color)
	for i in range(current_row + 1):
		sheet.row_dimensions[i + 1].height = 15
	for j in range(1, columns + 1):
		sheet.column_dimensions[get_column_letter(j)].width = greatest_column[j]
	if not excel_file.views:
		excel_file.views.append(openpyxl.workbook.views.BookView())
	excel_file.save('results.xlsx')

def build_styles():
	global excel_file
	from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment, Border, Side
	alignment = Alignment(
		horizontal = 'center',
		vertical = 'center',
	)
	border = Border(
		left = Side(style = 'thin'),
		right = Side(style = 'thin'),
		top = Side(style = 'thin'),
		bottom = Side(style = 'thin'),
	)
	top_style = NamedStyle('top_style')
	top_style.alignment = alignment
	top_style.border = border
	top_style.fill = PatternFill('solid', fgColor = "000000")
	top_style.font = Font(
		bold = True,
		color = 'FFFFFF',
		name = 'Calibri',
		size = 10,
	)
	excel_file.add_named_style(top_style)
	normal_style = NamedStyle('normal_style')
	normal_style.alignment = alignment
	normal_style.border = border
	normal_style.fill = PatternFill('solid', fgColor = "FFFFFF")
	normal_style.font = Font(
		color = '000000',
		name = 'Calibri',
		size = 10,
	)
	excel_file.add_named_style(normal_style)

class data:
	def __init__(self, regional, locale, station, cabinet, total, available, occupied):
		self.regional = regional
		self.locale = locale
		self.station = station
		self.cabinet = cabinet
		self.total = total
		self.available = available
		self.occupied = occupied
	def __lt__(self, other):
		if self.available != other.available:
			return self.available < other.available
		if self.regional != other.regional:
			return self.regional < other.regional
		if self.locale != other.locale:
			return self.locale < other.locale
		if self.station != other.station:
			return self.station < other.station
		return self.cabinet < other.cabinet
		
def get_cabinet(v):
	if len(v[ord('U') - ord('A')]) == 0:
		return v[ord('H') - ord('A')] + ' ARD RR'
	if v[ord('U') - ord('A')].find('ARD ') == -1:
		return v[ord('U') - ord('A')] + ' ARD ' + v[ord('V') - ord('A')]
	return v[ord('U') - ord('A')]

def main():
	current_file, previous_file, date_difference = read_config_file('config.txt')
	read_file(current_file)
	read_file(previous_file)
	build_database(current_file)
	build_excel_file(current_file, previous_file, date_difference)

def read_config_file(filename):
	global database
	database = {}
	try:
		with open(filename, 'r') as config_file:
			v = config_file.readlines()
			current_file = v[0].split('=')[1].strip().split('"')[1].strip()
			previous_file = v[1].split('=')[1].strip().split('"')[1].strip()
			date_difference = int(v[2].split('=')[1].strip().split('"')[1].strip())
			return (current_file, previous_file, date_difference)
	except Exception as e:
		print(e)
		print('Failed to read file: ' + filename)

def read_file(filename):
	global database
	database[filename] = {}
	try:
		with open(filename, 'r', encoding = 'ISO-8859-1') as input_file:
			technologies = [
				'huawei vdsl',
				'keymile vdsl',
			]
			occupied = [
				'auditoria',
				'ocupado',
				'reservado ngn',
			]
			available = [
				'disponivel',
				'disponivel ngn',
			]
			for line in input_file.readlines():
				v = line.split(';')
				technology = str(v[18]).strip().lower()
				if technology in technologies:
					status = str(v[4]).strip()
					regional = str(v[5]).strip()
					locale = str(v[6]).strip()
					station = str(v[7]).strip()
					cabinet = get_cabinet(v)
					port = +1 if status in available else (-1 if status in occupied else 0)
					add_port(filename, regional, locale, station, cabinet, port)
	except Exception as e:
		print(e)
		print('Failed to read file: ' + filename)

main()