from openpyxl import load_workbook

wbEXor = load_workbook('datasheets/Relatorio GPON - Ex-180125.xlsx')
wbCNor = load_workbook('datasheets/Relatorio GPON - Cn-180125.xlsx')
wsEx = wbEXor['RelatórioAtual']
wsCn = wbCNor['RelatórioAtual'] 


wbEXmy = load_workbook('datasheets/saidaEX.xlsx')
wbCNmy = load_workbook('datasheets/saidaCN.xlsx')
myEx = wbEXmy.active
myCn = wbCNmy.active



print("comparando Expansao...")
num_rows = max( wsEx.max_row,    myEx.max_row    )
num_cols = max( wsEx.max_column, myEx.max_column )

for i in range(1, num_rows+1):
    for j in range(1, num_cols+1):
        if wsEx.cell(row=i, column=j).value != myEx.cell(row=i, column=j).value:
            print("%s ||||| %s" %( wsEx.cell(row=i, column=j).value, myEx.cell(row=i, column=j).value))
            input()
            

print("Analise completa")

print("comparando Concessao...")
num_rows = max( wsCn.max_row,    myCn.max_row    )
num_cols = max( wsCn.max_column, myCn.max_column )

for i in range(1, num_rows+1):
    for j in range(1, num_cols+1):
        if wsCn.cell(row=i, column=j).value != myCn.cell(row=i, column=j).value:
            print("%s ||||| %s" %( wsCn.cell(row=i, column=j).value, myCn.cell(row=i, column=j).value))
            input()
            

print("Analise completa")



