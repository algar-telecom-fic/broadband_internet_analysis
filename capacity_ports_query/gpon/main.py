from aux_functions import now, date_difference

#this is the library we're using to interact with datasheets files
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook


#function to add the data from the file to the database
def add_port(concession_type, locale, station, cto, status):
    #using the global database
    global database
    
    #if any of these values is not in the dictionary yet, create the it's dictionary
    if locale not in database[concession_type]:
        database[concession_type][locale] = {}
    if station not in database[concession_type][locale]:
        database[concession_type][locale][station] = {}
        
    
    #the cto dictionary is where we're gonna count the records by their status 
    if cto not in database[concession_type][locale][station]:
        database[concession_type][locale][station][cto] = {
            'DEFEITO':   0,
            'DESIGNADO': 0,
            'OCUPADO':   0,
            'RESERVADO': 0,
            'VAGO':      0,
            'AUDITORIA': 0,
            'total':     0,
        }
    
    #all lines add in the total
    database[concession_type][locale][station][cto]['total'] += 1
    
    #each line also adds in it's own status
    database[concession_type][locale][station][cto][status] += 1

#function to manipulate everything in the database and ordenate it
def build_database(concession_type):
    global database

    #we're gonna use this list to store and sort the records
    v = []
    
    #iterates for everything
    for locale in database[concession_type]:
        for station in database[concession_type][locale]:
            for cto in database[concession_type][locale][station]:
                        #append each record to the empty list              
                        v.append(
                            #create a 'data' object with all the information
                            data(
                                locale,
                                station,
                                cto,
                                database[concession_type][locale][station][cto]['DEFEITO'],
                                database[concession_type][locale][station][cto]['DESIGNADO'],
                                database[concession_type][locale][station][cto]['OCUPADO'],
                                database[concession_type][locale][station][cto]['RESERVADO'],
                                database[concession_type][locale][station][cto]['VAGO'],
                                database[concession_type][locale][station][cto]['total'], 
                                #we are ommiting the 'AUDITORIA' because it doesn't appear in the final sheet      
                            )
                        )
    
    #after everything is in the list we can sort it
    v.sort()
    #we store our list in the same part of the database all the data was to use it globally afterwards
    database[concession_type] = v
 
#this function creates the xlsx file where we're gonna store the results
#we have one file for each type of concession
def build_excel_file(concession_type, concession_list):
    #using the global database and excel file
    global database, excel_file
    
    
    
    #'Workbook' is the main class of the library, it represents a datasheet file (xlsx)
    excel_file = Workbook()
    
    #first we create the styles we are going to use
    build_styles()
    
    #we're going to create several sheets within the same excel file
    #and we have a function for each of the sheets
    
    #this is the first one
    build_relatorioAtual(concession_type)
    #for the second one we need to first declare a new sheet
    excel_file.create_sheet()
    #then we can create the second one
    build_PortaCTOE(concession_type)
    
    excel_file.create_sheet()
    
    build_Crescimento(concession_type, concession_list)
    
    #depending of the concession type we choose a different name for the result file
    if concession_type == 'concession':
        filename = 'datasheets/resultsCN.xlsx'
    else:
        filename = 'datasheets/resultsEX.xlsx'
     
       
    #print(filename)
    #in the end, we save the file with the name whe chose
    excel_file.save(filename)

#this function build the styles we're going to use in the files
def build_styles():
    #the excel file we are editating now is a global variable
    global excel_file
    

    #this is for aligning the text inside the cells
    alignment = Alignment(
        horizontal = 'center',
        vertical = 'center',
    )
    
    #aplying a border to the cell
    border = Border(
		left = Side(style = 'thin'),
		right = Side(style = 'thin'),
		top = Side(style = 'thin'),
		bottom = Side(style = 'thin'),
	)
    
    #the default font style used in the headlines
    font = Font(
        #this makes the font bold for the headlines
        bold = True,
        color = '000000',
        name = 'Calibri',
        size = 11,
    )
    
    #creating the top style for the headline cells
    top_style = NamedStyle('top_style')
    #applying our configuration to the style we just created
    top_style.alignment = alignment
    top_style.font = font
    top_style.border = border
    #adding the style to the file
    excel_file.add_named_style(top_style)

    #the normal style
    normal_style = NamedStyle('normal_style')
    #in this style we don't need an aligment nor a font, we're gonna use the default
    #which is:
        #horizontal = 'left',
        #vertical = 'bottom',
        #fontName = 'Calibri',
        #fontSize = 11,
        #color = '000000'
        
    normal_style.border = border
    
    #adding the style to the file
    excel_file.add_named_style(normal_style)
    
    
    #this is style is the same as the normal but with center aligning 
    center_style = NamedStyle('center_style')
    center_style.alignment = alignment
    center_style.border = border
    excel_file.add_named_style(center_style)


#function for creating one sheet of the excel file
def build_relatorioAtual(concession_type):
    
    
    #here we select the last sheet of the file as the one we're going to use
    sheet = excel_file.worksheets[-1]
    #changing the title of the sheet
    sheet.title = 'RelatórioAtual'
    
    #this particular sheet has 10 columns
    num_columns = 10
    
    #filling the first row of cells with the headlines values
    sheet.cell(row = 1, column = 1).value = 'LOCALIDADE'
    sheet.cell(row = 1, column = 2).value = 'ESTACAO'
    sheet.cell(row = 1, column = 3).value = 'CTO'
    sheet.cell(row = 1, column = 4).value = 'DEFEITO'
    sheet.cell(row = 1, column = 5).value = 'DESIGNADO'
    sheet.cell(row = 1, column = 6).value = 'OCUPADO'
    sheet.cell(row = 1, column = 7).value = 'RESERVADO'
    sheet.cell(row = 1, column = 8).value = 'VAGO'
    sheet.cell(row = 1, column = 9).value = 'Total Geral'
    
    
    #this is to change the color of the cells from the first row
    for i in range(1, num_columns+1):
        #the .fill property controls the color
        # 'D9E1F2' is the hex representation of a light, clear, beautiful blue
        sheet.cell(row = 1, column = i).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid') 
    
    #now we're going to fill the cells with the data collected
    #starting from the second row
    current_row = 2
    
    #iterating for the records in the database
    #each record is a 'data' object
    for record in database[concession_type]:
        #in each of the nine cells in the current row we fill in the right information
        #these cells don't need any style, they look OK with the default
        sheet.cell(row = current_row, column = 1).value = record.locale
        sheet.cell(row = current_row, column = 2).value = record.station
        sheet.cell(row = current_row, column = 3).value = record.cto
        sheet.cell(row = current_row, column = 4).value = record.defeito
        sheet.cell(row = current_row, column = 5).value = record.designado
        sheet.cell(row = current_row, column = 6).value = record.ocupado
        sheet.cell(row = current_row, column = 7).value = record.reservado
        sheet.cell(row = current_row, column = 8).value = record.vago
        sheet.cell(row = current_row, column = 9).value = record.total
        #atualize the current row
        current_row+=1
    
    #we need to save the date we generated this file
    sheet.cell(row = 1, column = 10).value = 'Data'
    sheet.cell(row = 2, column = 10).value = now()
    

#function for creating the other sheet of the excel file
def build_PortaCTOE(concession_type):

    #using the last worksheet created
    sheet = excel_file.worksheets[-1]
    sheet.title = '1-Porta CTOE'
    
    #this time we have 7 columns
    num_columns = 7
    #these are the headlines
    #this time the headlines are two rows thick
    
    sheet.cell(row = 1, column = 1).value = 'LOCALIDADE'
    sheet.cell(row = 1, column = 2).value = 'ESTACAO'
    sheet.cell(row = 1, column = 3).value = 'CTO'
    sheet.cell(row = 1, column = 4).value = 'Possibilidade de Vendas'
    sheet.cell(row = 1, column = 5).value = 'Capacidade CTOE'
    
    #these are the headlines of the second row
    sheet.cell(row = 2, column = 5).value = 'Ocupado'
    sheet.cell(row = 2, column = 6).value = 'Disponível'
    sheet.cell(row = 2, column = 7).value = 'Instalado'
    
    #this is for merging the cells that ocuppy more than one row 
    #it mantains the content of the left-most upper-most cell
    sheet.merge_cells('A1:A2')
    sheet.merge_cells('B1:B2')
    sheet.merge_cells('C1:C2')
    sheet.merge_cells('D1:D2')
    sheet.merge_cells('E1:G1')
    
    #this apply the same top_style we created before to all cells of the first and second row
    #the cells that were merged receive the style two times
    for i in range(1, num_columns+1):
        sheet.cell(row = 1, column = i).style = 'top_style'
        sheet.cell(row = 2, column = i).style = 'top_style'
        
    
    #setting the color of the blue cells
    sheet.cell(row = 1, column = 1).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    sheet.cell(row = 1, column = 2).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    #setting the color of the yellow cells
    sheet.cell(row = 1, column = 3).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') 
    sheet.cell(row = 1, column = 4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    #setting the color of the green cells
    sheet.cell(row = 1, column = 5).fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
    
    #setting the color of the orange-ish cells
    sheet.cell(row = 2, column = 5).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    sheet.cell(row = 2, column = 6).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    sheet.cell(row = 2, column = 7).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    
    
    #starting for the third row this time
    current_row = 3
    #iterate for all the records
    for record in database[concession_type]:
    
        #each cell receive content and style
        sheet.cell(row = current_row, column = 1).value = record.locale
        sheet.cell(row = current_row, column = 1).style = 'normal_style'
        
        sheet.cell(row = current_row, column = 2).value = record.station
        sheet.cell(row = current_row, column = 2).style = 'normal_style'
        
        sheet.cell(row = current_row, column = 3).value = record.cto
        sheet.cell(row = current_row, column = 3).style = 'normal_style'
        
        #the content of this one is conditional, depending of other property
        sheet.cell(row = current_row, column = 4).value = ("Sim - %s" % record.vago) if (int(record.vago) > 0) else "Não - Indisponibilidade CTOE"
        sheet.cell(row = current_row, column = 4).style = 'center_style'
        
        #the content of this other one is the sum of two properties
        sheet.cell(row = current_row, column = 5).value = int(record.designado) + int(record.ocupado)
        sheet.cell(row = current_row, column = 5).style = 'center_style'
        
        sheet.cell(row = current_row, column = 6).value = record.vago
        sheet.cell(row = current_row, column = 6).style = 'center_style'
        
        sheet.cell(row = current_row, column = 7).value = record.total
        sheet.cell(row = current_row, column = 7).style = 'center_style'
        
        current_row+=1
        


def build_Crescimento(concession_type, concession_list):
    #using the last worksheet created
    sheet = excel_file.worksheets[-1]
    sheet.title = '3-Tx-Crescimento'
    
    sheet.cell(row = 1, column = 1).value = 'LOCALIDADE'
    sheet.cell(row = 1, column = 2).value = 'Ocupação Atual'
    sheet.cell(row = 1, column = 3).value = 'Ocupação Anterior'
    sheet.cell(row = 1, column = 4).value = 'Tx Crescimento (Mês)'
    sheet.cell(row = 1, column = 5).value = 'Capacidade Atual'
    sheet.cell(row = 1, column = 6).value = 'Espectativa de Esgotamento'
    sheet.cell(row = 1, column = 7).value = 'Visão de Capacidade'
    
    for i in range(1, 8):
        sheet.cell(row = 1, column = i).style = 'top_style'
        
    sheet.cell(row = 1, column = 1).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    sheet.cell(row = 1, column = 2).fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
    sheet.cell(row = 1, column = 3).fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
    
    sheet.cell(row = 1, column = 4).fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
    sheet.cell(row = 1, column = 5).fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
    sheet.cell(row = 1, column = 6).fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
    sheet.cell(row = 1, column = 7).fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
    
    
    current_row = 2
    
    olddata = 'old_'+concession_type
    for locale in concession_list:
        ocupado = 0
        
        if locale in database[olddata]:
            for station in database[olddata][locale]:
                for cto in database[olddata][locale][station]:
                    ocupado+=database[olddata][locale][station][cto]['OCUPADO']
        
        sheet.cell(row = current_row, column = 3).value = ocupado
        current_row+=1
    
    
    current_locale = database[concession_type][0].locale

    ocupado = 0
    vago = 0
    current_row = 2

    for record in database[concession_type]:
        if record.locale != current_locale:
            
            sheet.cell(row = current_row, column = 1).value = current_locale
            sheet.cell(row = current_row, column = 2).value = ocupado
            sheet.cell(row = current_row, column = 5).value = vago
            
            current_locale = record.locale
            current_row+=1
            ocupado = 0
            vago = 0
        
        ocupado+=record.ocupado
        vago+=record.vago
    
    sheet.cell(row = current_row, column = 1).value = current_locale
    sheet.cell(row = current_row, column = 2).value = ocupado
    sheet.cell(row = current_row, column = 5).value = vago
        
    global old_date
    
    days = date_difference(old_date)
    
    for i in range(2, current_row+1):
        deltaOcupacao =  int(sheet.cell(row = i, column = 2).value) - int(sheet.cell(row = i, column = 3).value)
        tx_crescimento = ((deltaOcupacao / days) *30)
        sheet.cell(row = i, column = 4).value = "%.2f" % tx_crescimento
        
        sheet.cell(row = i, column = 1).style = 'normal_style'
        sheet.cell(row = i, column = 2).style = 'normal_style'
        sheet.cell(row = i, column = 3).style = 'normal_style'
        sheet.cell(row = i, column = 4).style = 'center_style'
        sheet.cell(row = i, column = 5).style = 'center_style'
        sheet.cell(row = i, column = 6).style = 'center_style'
        sheet.cell(row = i, column = 7).style = 'center_style'
        
        try:
            month_prevision = (float(sheet.cell(row = i, column = 5).value) / tx_crescimento)
            sheet.cell(row = i, column = 6).value = "%.2f" % month_prevision
            
            if month_prevision < 0.00:
                sheet.cell(row = i, column = 7).value = 'Decrescimento'
            if 0.01 < month_prevision <= 1.00:
                sheet.cell(row = i, column = 7).value = '1 - Esgota até Um Mês'
            if 1.00 < month_prevision <= 2.00:
                sheet.cell(row = i, column = 7).value = '2 - Esgota até Dois Meses'
            if 2.00 < month_prevision <= 3.00:
                sheet.cell(row = i, column = 7).value = '3 - Esgota até Três Meses'
            if 3.00 < month_prevision <= 4.00:
                sheet.cell(row = i, column = 7).value = '4 - Esgota até Quatro Meses'
            if 4.00 < month_prevision <= 5.00:
                sheet.cell(row = i, column = 7).value = '5 - Esgota até Cinco Meses'
            if 5.00 < month_prevision <= 6.00:
                sheet.cell(row = i, column = 7).value = '6 - Esgota até Seis Meses'
            if 6.00 < month_prevision <= 7.00:
                sheet.cell(row = i, column = 7).value = '7 - Esgota até Sete Meses'
            if 7.00 < month_prevision <= 8.00:
                sheet.cell(row = i, column = 7).value = '8 - Esgota até Oito Meses'
            if 8.00 < month_prevision <= 9.00:
                sheet.cell(row = i, column = 7).value = '9 - Esgota até Nove Meses'
            if 9.00 < month_prevision <= 10.00:
                sheet.cell(row = i, column = 7).value = '10 - Esgota até Dez Meses'
            if 10.00 < month_prevision:
                sheet.cell(row = i, column = 7).value = '11 - Esgota em Mais de 10 Meses'
             
        except ZeroDivisionError as e:
            print(e)
            sheet.cell(row = i, column = 6).value = "Indisponível"
            sheet.cell(row = i, column = 7).value = '11 - Esgota em Mais de 10 Meses'
    


#this class just wrap all the usefull information
#it serves mostly just to ordenate the results
class data:
    #the constructor is pretty default
    def __init__(self, locale, station, cto, defeito, designado, ocupado, reservado, vago, total):
        self.locale = locale
        self.station = station
        self.cto = cto
        self.defeito = defeito
        self.designado = designado
        self.ocupado = ocupado
        self.reservado = reservado
        self.vago = vago
        self.total = total
    
    #this is the '<' operator 
    def __lt__(self, other):
        if self.locale != other.locale:
            return self.locale < other.locale
        if self.station != other.station:
            return self.station < other.station
        
        #we're sure the cto is unique so we don't need to compare nothing beyond it
        return self.cto < other.cto
    
    #this is for converting the class to a string when we need to debug the code
    def __repr__(self):
        return "%s,%s,%s,%s,%s,%s,%s,%s,%s,"%(
            self.locale,
            self.station,
            self.cto,
            self.defeito,
            self.designado,
            self.ocupado,
            self.reservado,
            self.vago,
            self.total,
        )
  
#function for extrating the useful data from the file
def read_file(filename, concession, expansion):
    
    #declaring the database as global will help we use it across different functions 
    global database
    
    #we're spliting the database in two areas to separate the two types of locales
    database['concession'] = {}
    database['expansion'] = {}
    
    #open the file
    with open(filename, 'r',  encoding = 'ISO-8859-1') as input_file:

        #iterate for all the lines
        for line in input_file.readlines():

            #split the csv by the semi-colon
            v = line.split(';')
            
            #take the cto status form the list
            status_cto = str(v[4]).strip()
            #we must consider only the existing CTO
            if status_cto == "EXISTENTE":
                #separate the important properties
                locale = str(v[14]).strip()
                station = str(v[15]).strip()
                cto = str(v[1]).strip()     
                status = str(v[13]).strip()    
                
                #printing in the screen just for debug purposes
                #print("%s %s %s %s" %(locale, station, cto, status))
                
                if locale in concession:
                    #add this line's data to the database
                    add_port('concession', locale, station, cto, status)
                elif locale in expansion:
                    add_port('expansion', locale, station, cto, status)
            


def read_excel_file(filename, concession_type):
    global database, old_date
    
    wb = load_workbook(filename)
    ws = wb['RelatórioAtual']
    
    num_rows = ws.max_row 
    
    database[concession_type] = {}
    
    for i in range(2, num_rows+1):
        locale  = ws.cell(row = i, column = 1).value
        station = ws.cell(row = i, column = 2).value
        cto     = ws.cell(row = i, column = 3).value
        
        if locale not in database[concession_type]:
            database[concession_type][locale] = {}
        if station not in database[concession_type][locale]:
            database[concession_type][locale][station] = {}
        
        database[concession_type][locale][station][cto] = {
            'DEFEITO':   ws.cell(row = i, column = 4).value,
            'DESIGNADO': ws.cell(row = i, column = 5).value,
            'OCUPADO':   ws.cell(row = i, column = 6).value,
            'RESERVADO': ws.cell(row = i, column = 7).value,
            'VAGO':      ws.cell(row = i, column = 8).value,
            'total':     ws.cell(row = i, column = 9).value,
        }
    
    old_date = str(ws.cell(row = 2, column = 10).value)
    print("aquiii %s" %old_date)

#this file contains the relation between the cities and each concession type
def read_concession_file(filename):


    #loading the file from the disc
    wb = load_workbook(filename)
    #selectin the right sheet
    ws = wb['todas_localidades_existentes']

    #both lists start empty
    concession = []
    expansion  = []

    #getting the max number of rows
    num_rows = ws.max_row
    
    #iterating for all the rows of the sheet
    for i in range(2, num_rows+1):
        #separating the city and the concession type
        locale = ws.cell(row=i, column=1).value
        locale_type = ws.cell(row=i, column=2).value 
        
        #depending of the type, we append the locale to the right list
        if locale_type == 'CONCESSÃO':
            concession.append(locale)
        else:
            expansion.append(locale)
    
    #returning the two lists
    return (concession, expansion)


def main():
    global database
    database = {}

    concession_filename = 'datasheets/CNxEX_MOLDE.xlsx'
    concession_lists = read_concession_file(concession_filename)

    filename='datasheets/Circuitos CTO-01-25.csv'
    read_file(filename, concession_lists[0], concession_lists[1])

    oldCn = 'datasheets/oldCn.xlsx'
    oldEx = 'datasheets/oldEx.xlsx'

    read_excel_file(oldCn, 'old_concession')
    read_excel_file(oldEx, 'old_expansion')


    build_database('concession')
    build_database('expansion')

    


    build_excel_file('concession', concession_lists[0])
    build_excel_file('expansion', concession_lists[1])



main()




"""
PERGUNTAR SOBRE:
    AUDITORIA
    DIFERENÇA DE LINHAS ENTRE RELATORIO ATUAL E PORTA CTOE EXPANSAO
    PORTA CTOE: CTOE.ocupado = relatorio.ocupado + relatorio.designado
"""






