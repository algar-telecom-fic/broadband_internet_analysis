from openpyxl import Workbook
from openpyxl.styles import PatternFill

newbook = Workbook()

sheet = newbook.worksheets[0]

sheet.merge_cells('A1:A2')

sheet.merge_cells('D2:E4')

sheet.cell(row = 1, column = 1).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

newbook.save('datasheets/teste.xlsx')

