def build_excel_file(current_file, previous_file, date_difference):
	global database, excel_file
	from math import ceil
	from openpyxl import Workbook
	from openpyxl.utils import get_column_letter
	from openpyxl.styles import PatternFill
	excel_file = Workbook()
	sheet = excel_file.worksheets[0]
	sheet.title = 'Results'
	build_styles()
	current_row = 1
	columns = 10
	greatest_column = [0 for i in range(columns + 1)]
	sheet.cell(row = current_row, column = 1).value = 'Node'
	sheet.cell(row = current_row, column = 2).value = 'CMTS'
	sheet.cell(row = current_row, column = 3).value = 'Qtd. Clientes'
	sheet.cell(row = current_row, column = 4).value = 'Capacidade (Mbps)'
	sheet.cell(row = current_row, column = 5).value = 'Utilização média entre as portadoras (Mbps)'
	sheet.cell(row = current_row, column = 6).value = 'Utilização média entre as portadoras (%)'
	sheet.cell(row = current_row, column = 7).value = 'Crescimento mensal (%)'
	sheet.cell(row = current_row, column = 8).value = 'Crescimento mensal (Mbps)'
	sheet.cell(row = current_row, column = 9).value = 'Previsão de congestionamento (Mês)'
	sheet.cell(row = current_row, column = 10).value = 'Previsão de congestionamento'
	for j in range(1, columns + 1):
		sheet.cell(row = current_row, column = j).style = 'top_style'
		greatest_column[j] = max(greatest_column[j], len(str(sheet.cell(row = current_row, column = j).value)))
	for i in database[current_file]:
		current_row += 1
		sheet.cell(row = current_row, column = 1).value = i['node']
		sheet.cell(row = current_row, column = 2).value = '?'
		sheet.cell(row = current_row, column = 3).value = '?'
		sheet.cell(row = current_row, column = 4).value = i['capacity']
		sheet.cell(row = current_row, column = 5).value = i['capacity'] * i['usage'] / 100.0
		sheet.cell(row = current_row, column = 6).value = i['usage']
		sheet.cell(row = current_row, column = 7).value = '?'
		sheet.cell(row = current_row, column = 8).value = '?'
		sheet.cell(row = current_row, column = 9).value = '?'
		sheet.cell(row = current_row, column = 10).value = '?'
		# ~ try:
			# ~ before = database[previous_file][i.regional][i.locale][i.station][i.cabinet]['occupied']
			# ~ median = (i.occupied - before) / (date_difference / 30)
			# ~ sheet.cell(row = current_row, column = 8).value = round(median, 1)
			# ~ if i.available == 0:
				# ~ sheet.cell(row = current_row, column = 9).value = 'Esgotado'
			# ~ elif median > 0:
				# ~ value = ceil(i.available / median)
				# ~ if value == 1:
					# ~ sheet.cell(row = current_row, column = 9).value = 'Esgota em até 1 mês'
				# ~ elif value > 10:
					# ~ sheet.cell(row = current_row, column = 9).value = 'Esgota em mais de 10 meses'
				# ~ else:
					# ~ sheet.cell(row = current_row, column = 9).value = 'Esgota em até ' + str(value) + ' meses'
			# ~ elif median < 0:
				# ~ sheet.cell(row = current_row, column = columns).value = 'Decrescimento'
			# ~ else:
				# ~ sheet.cell(row = current_row, column = columns).value = 'Estável'
		# ~ except Exception as e:
			# ~ print(e)
			# ~ sheet.cell(row = current_row, column = 8).value = 'Sem histórico'
			# ~ sheet.cell(row = current_row, column = 9).value = 'Esgotado' if i.available == 0 else 'Sem histórico'
		for j in range(1, columns + 1):
			sheet.cell(row = current_row, column = j).style = 'normal_style'
			greatest_column[j] = max(greatest_column[j], len(str(sheet.cell(row = current_row, column = j).value)))
		# ~ current_color = 'FF0000' if i.available == 0 else ('00FF00' if i.available > 10 else 'FFFF00')
		# ~ sheet.cell(row = current_row, column = 6).fill = PatternFill('solid', fgColor = current_color)
	for i in range(current_row + 1):
		sheet.row_dimensions[i + 1].height = 15
	for j in range(1, columns + 1):
		sheet.column_dimensions[get_column_letter(j)].width = greatest_column[j]
	if not excel_file.views:
		excel_file.views.append(openpyxl.workbook.views.BookView())
	excel_file.save('results.xlsx')

def build_styles():
	global excel_file
	from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment, Border, Side
	alignment = Alignment(
		horizontal = 'center',
		vertical = 'center',
	)
	border = Border(
		left = Side(style = 'thin'),
		right = Side(style = 'thin'),
		top = Side(style = 'thin'),
		bottom = Side(style = 'thin'),
	)
	top_style = NamedStyle('top_style')
	top_style.alignment = alignment
	top_style.border = border
	top_style.fill = PatternFill('solid', fgColor = "000000")
	top_style.font = Font(
		bold = True,
		color = 'FFFFFF',
		name = 'Calibri',
		size = 10,
	)
	excel_file.add_named_style(top_style)
	normal_style = NamedStyle('normal_style')
	normal_style.alignment = alignment
	normal_style.border = border
	normal_style.fill = PatternFill('solid', fgColor = "FFFFFF")
	normal_style.font = Font(
		color = '000000',
		name = 'Calibri',
		size = 10,
	)
	excel_file.add_named_style(normal_style)

def main():
	current_file, previous_file, date_difference = read_config_file('config.txt')
	read_file(current_file)
	read_file(previous_file)
	build_excel_file(current_file, previous_file, date_difference)

def read_config_file(filename):
	global database
	database = {}
	try:
		with open(filename, 'r') as config_file:
			v = config_file.readlines()
			current_file = v[0].split('=')[1].strip().split('"')[1].strip()
			previous_file = v[1].split('=')[1].strip().split('"')[1].strip()
			date_difference = int(v[2].split('=')[1].strip().split('"')[1].strip())
			return (current_file, previous_file, date_difference)
	except Exception as e:
		print(e)
		print('Failed to read file: ' + filename)

def read_file(filename):
	global database
	database[filename] = []
	try:
		with open(filename, 'r', encoding = 'ISO-8859-1') as input_file:
			lines = input_file.readlines()
			for i in range(len(lines)):
				lines[i] = lines[i].strip()
			idx = -1
			while idx + 1 < len(lines):
				idx += 1
				if len(lines[idx]) == 0:
					continue
				node_name = lines[idx]
				qtd_interfaces = 0
				total_capacity = 0
				total_usage = 0
				idx += 2
				while idx + 1 < len(lines):
					idx += 1
					if len(lines[idx]) == 0:
						break
					qtd_interfaces += 1
					v = lines[idx].split(';')
					total_capacity += int(v[2])
					total_usage += float(v[6])
				median_usage = total_usage / qtd_interfaces
				database[filename].append({
					'node': node_name,
					'capacity': total_capacity,
					'usage': median_usage
				})
	except Exception as e:
		print(e)
		print('Failed to read file: ' + filename)

main()