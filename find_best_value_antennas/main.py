import openpyxl, math
from copy import copy

info = {}
area = []
links = []
sheets_file = 0
wind_area_constants = {}
modulation_constants = []
last_channel_constants = []
red_color = openpyxl.styles.PatternFill(start_color = 'FFFF0000', end_color = 'FFFF0000', fill_type = 'solid')
white_color = openpyxl.styles.PatternFill(start_color = 'FFFFFFFF', end_color = 'FFFFFFFF', fill_type = 'solid')

def add_result_line(idx):
	idx += 1
	global sheets_file, white_color
	sheet = sheets_file.worksheets[2]
	for j in range(11):
		cell = sheet.cell(row = idx - 1, column = j + 1)
		sheet.cell(row = idx, column = j + 1).font = copy(cell.font)
		sheet.cell(row = idx, column = j + 1).border = copy(cell.border)
		sheet.cell(row = idx, column = j + 1).fill = white_color if j > 0 else copy(cell.fill)
		sheet.cell(row = idx, column = j + 1).number_format = copy(cell.number_format)
		sheet.cell(row = idx, column = j + 1).protection = copy(cell.protection)
		sheet.cell(row = idx, column = j + 1).alignment = copy(cell.alignment)
		sheet.cell(row = idx, column = j + 1).value = '-'
	return idx

def build_constants():
	#  5 GHz and modulation QAM -> 128, 256, 512, 1024
	last_channel_constants.append(5090)
	modulation_constants.append([27, 23, 21, 17])
	#  6 GHz - LOW and modulation QAM -> 128, 256, 512, 1024
	last_channel_constants.append(6405)
	modulation_constants.append([27, 23, 21, 17])
	#  6 GHz - UP and modulation QAM -> 128, 256, 512, 1024
	last_channel_constants.append(7080)
	modulation_constants.append([27, 23, 21, 17])
	#  7 GHz and modulation QAM -> 128, 256, 512, 1024
	last_channel_constants.append(7708)
	modulation_constants.append([25, 21, 19, 15])
	#  8 GHz and modulation QAM -> 128, 256, 512, 1024
	last_channel_constants.append(8267)
	modulation_constants.append([25, 21, 19, 15])
	#  8,5 GHz and modulation QAM -> 128, 256, 512, 1024
	last_channel_constants.append(8482)
	modulation_constants.append([25, 21, 19, 15])
	#  15 GHz and modulation QAM -> 128, 256, 512, 1024
	last_channel_constants.append(15327)
	modulation_constants.append([29, 25, 23, 19])
	#  18 GHz and modulation QAM -> 128, 256, 512, 1024
	last_channel_constants.append(19672)
	modulation_constants.append([29, 25, 23, 19])
	#  23 GHz and modulation QAM -> 128, 256, 512, 1024
	last_channel_constants.append(23562)
	modulation_constants.append([29, 25, 23, 19])
	#  global wind_area_constants
	wind_area_constants[0.3] = 0.11
	wind_area_constants[0.6] = 0.45
	wind_area_constants[0.9] = 1.02
	wind_area_constants[1.0] = 1.26
	wind_area_constants[1.2] = 1.81
	wind_area_constants[1.8] = 4.07
	wind_area_constants[2.4] = 7.23
	wind_area_constants[3.0] = 11.30
	wind_area_constants[3.6] = 16.28
	wind_area_constants[3.7] = 17.19

def open_sheets_file():
	global sheets_file
	sheets_file = openpyxl.load_workbook('input.xlsx')

def read_info():
	global sheets_file, red_color
	sheet = sheets_file.worksheets[0]
	idx = 0
	while True:
		current_frequency = sheet.cell(row = 6, column = idx * 4 + 1).value
		if current_frequency == None:
			break
		info[current_frequency] = {}
		info[current_frequency]['channel'] = last_channel_constants[idx]
		info[current_frequency]['modulation'] = modulation_constants[idx]
		#  antenna
		info[current_frequency]['antenna'] = {}
		before_gain = float(-1000000000000)
		for i in range(9, 19):
			diameter = sheet.cell(row = i, column = (idx * 4) + 1).value
			gain = sheet.cell(row = i, column = (idx * 4) + 3).value
			if diameter != None and gain != None:
				try:
					gain = float(gain)
				except Exception:
					sheet.cell(row = i, column = (idx * 4) + 3).fill = red_color
					continue
				info[current_frequency]['antenna'][diameter] = gain
				if gain < before_gain:
					sheet.cell(row = i, column = (idx * 4) + 3).fill = red_color
				else:
					before_gain = gain
		#  odu
		before_transmission = float(1000000000000)
		before_reception = float(-1000000000000)
		info[current_frequency]['odu'] = {}
		for i in range(22, 26):
			modulation = sheet.cell(row = i, column = (idx * 4) + 1).value
			transmission = sheet.cell(row = i, column = (idx * 4) + 2).value
			reception = sheet.cell(row = i, column = (idx * 4) + 3).value
			if modulation != None and transmission != None and reception != None:
				try:
					transmission = float(transmission)
				except Exception:
					sheet.cell(row = i, column = (idx * 4) + 2).fill = red_color
					continue
				try:
					reception = float(reception)
				except Exception:
					sheet.cell(row = i, column = (idx * 4) + 3).fill = red_color
					continue
				info[current_frequency]['odu'][modulation] = (transmission, reception)
				if transmission > before_transmission:
					sheet.cell(row = i, column = (idx * 4) + 2).fill = red_color
				else:
					before_transmission = transmission
				if reception < before_reception:
					sheet.cell(row = i, column = (idx * 4) + 3).fill = red_color
				else:
					before_reception = reception
		#  splitter
		info[current_frequency]['splitter'] = {}
		primary_loss = sheet.cell(row = 27, column = (idx * 4) + 3).value
		secondary_loss = sheet.cell(row = 28, column = (idx * 4) + 3).value
		if primary_loss != None:
			try:
				primary_loss = float(primary_loss)
			except Exception:
				sheet.cell(row = 27, column = (idx * 4) + 3).fill = red_color
			info[current_frequency]['splitter']['primary'] = primary_loss
		if secondary_loss != None:
			try:
				secondary_loss = float(secondary_loss)
			except Exception:
				sheet.cell(row = 28, column = (idx * 4) + 3).fill = red_color
			info[current_frequency]['splitter']['secondary'] = secondary_loss
		idx += 1

def read_links():
	global sheets_file
	sheet = sheets_file.worksheets[1]
	i = 2
	while True:
		vertex_A = sheet.cell(row = i, column = 1).value
		vertex_B = sheet.cell(row = i, column = 2).value
		distance = sheet.cell(row = i, column = 7).value
		if vertex_A == None or vertex_B == None or distance == None:
			break
		links.append((float(distance), vertex_A, vertex_B))
		i += 1

def save():
	global sheets_file
	if not sheets_file.views:
		sheets_file.views.append(openpyxl.workbook.views.BookView())
	sheets_file.save('output.xlsx')

def solve():
	global sheets_file
	sheet = sheets_file.worksheets[2]
	output_idx = 1
	link_flag = True
	for link in links:
		if link_flag:
			link_flag = False
			output_idx = add_result_line(output_idx)
		frequency_flag = False
		for frequency in info:
			if frequency_flag:
				frequency_flag = False
				output_idx = add_result_line(output_idx)
			channel = info[frequency]['channel']
			if channel > 8482 and link[0] > 10.0:
				continue
			if channel > 19672 and link[0] > 5.0:
				continue
			modulation_idx = -1
			for modulation in info[frequency]['odu']:
				modulation_idx += 1
				best_wind_area = 1123456789.0
				for diameter_A in info[frequency]['antenna']:
					gain_A = info[frequency]['antenna'][diameter_A]
					for diameter_B in info[frequency]['antenna']:
						gain_B = info[frequency]['antenna'][diameter_B]
						if 'primary' not in info[frequency]['splitter']:
							continue
						transmission = info[frequency]['odu'][modulation][0]
						reception = info[frequency]['odu'][modulation][1]
						at1 = info[frequency]['splitter']['primary']
						at2 = info[frequency]['splitter']['primary']
						pathloss = 32.45 + 20.0 * (math.log10(link[0]) + math.log10(channel))
						signal = transmission - at1 + gain_A - pathloss + gain_B - at2
						received = math.ceil(signal - reception)
						wind_area = wind_area_constants[diameter_A] + wind_area_constants[diameter_B]
						# ~ print(str(link) + ' ' + str(frequency) + ' ' + str(modulation) + ' -> ' + '(' + str(diameter_A) + ', ' + str(diameter_B) + ')' + ' = ' + str(signal) + ' -> ' + 'valor(' + str(received) + ') ' + ('<' if received < info[frequency]['modulation'][modulation_idx] else '>=') + ' constante(' + str(info[frequency]['modulation'][modulation_idx]) + '); ' + 'area: ' + str(wind_area))
						if received < info[frequency]['modulation'][modulation_idx] or wind_area >= best_wind_area:
							continue
						if frequency_flag == False:
							sheet.cell(row = output_idx, column = 1).value = frequency
							sheet.cell(row = output_idx, column = 2).value = link[1]
							sheet.cell(row = output_idx, column = 3).value = link[2]
						best_wind_area = wind_area
						link_flag = frequency_flag = True
						sheet.cell(row = output_idx, column = 5 + (modulation_idx << 1)).value = best_wind_area
						sheet.cell(row = output_idx, column = 4 + (modulation_idx << 1)).value = '(' + str(diameter_A) + ', ' + str(diameter_B) + ')'

def main():
	build_constants()
	open_sheets_file()
	read_info()
	read_links()
	solve()
	save()

main()